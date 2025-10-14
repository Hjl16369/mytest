import streamlit as st
import pandas as pd
import os
import zipfile
from io import BytesIO
import shutil

st.set_page_config(page_title="工作日报统计工具", layout="centered")
st.title("📊 工作日报周统计工具")

st.markdown("""
该工具用于统计**开发与测试人员**一周的工作量。  
请上传一个包含多个人员日报的 **ZIP 压缩包**（每个日报为 `.xlsx` 文件）。  
""")

# === Step 1: 上传压缩包 ===
uploaded_file = st.file_uploader("📂 上传工作日报压缩包（ZIP 格式）", type=["zip"])

if uploaded_file is not None:
    # 临时解压目录
    temp_dir = "temp_daily_reports"
    
    # 清理旧的临时目录
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
    os.makedirs(temp_dir, exist_ok=True)

    # 解压文件
    try:
        with zipfile.ZipFile(uploaded_file, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        st.success("✅ 压缩包已上传并解压成功！")
    except Exception as e:
        st.error(f"❌ 解压失败：{e}")
        st.stop()

    # === Step 2: 点击开始处理 ===
    if st.button("🚀 开始处理日报"):
        # 使用 session_state 保存处理结果，避免下载后数据丢失
        st.session_state.processing = True
        st.session_state.all_records = []
        st.session_state.debug_info = []
    
    # 如果已经处理过，显示结果
    if hasattr(st.session_state, 'processing') and st.session_state.processing:
        all_records = st.session_state.all_records
        debug_info = st.session_state.debug_info

        def read_daily_report(file_path: str):
            """读取单个日报文件（xlsx），返回 DataFrame"""
            file_name = os.path.basename(file_path)
            
            # 跳过临时文件和隐藏文件
            if file_name.startswith('~$') or file_name.startswith('.'):
                return pd.DataFrame()
            
            try:
                excel_file = pd.ExcelFile(file_path)
                file_records = []
                
                for sheet in excel_file.sheet_names:
                    try:
                        # 读取整个工作表，不指定header
                        df_full = pd.read_excel(file_path, sheet_name=sheet, header=None)
                        
                        # 跳过空表
                        if df_full.empty or len(df_full) < 5:
                            st.warning(f"⚠️ {file_name} 的 sheet《{sheet}》数据不足，已跳过。")
                            continue
                        
                        # 读取人员信息（B2单元格，索引为[1,1]）
                        try:
                            person_name = str(df_full.iloc[1, 1]).strip()
                            if pd.isna(df_full.iloc[1, 1]) or person_name == '' or person_name == 'nan':
                                person_name = file_name.replace(".xlsx", "").replace(".xls", "")
                                st.info(f"ℹ️ {file_name} 的 sheet《{sheet}》B2单元格为空，使用文件名作为人员名")
                        except:
                            person_name = file_name.replace(".xlsx", "").replace(".xls", "")
                            st.warning(f"⚠️ {file_name} 的 sheet《{sheet}》无法读取B2单元格，使用文件名作为人员名")
                        
                        # 读取日期信息（B3单元格，索引为[2,1]）
                        try:
                            date_value = df_full.iloc[2, 1]
                            if pd.isna(date_value) or str(date_value).strip() == '':
                                date_str = sheet  # 使用sheet名称作为日期
                                st.info(f"ℹ️ {file_name} 的 sheet《{sheet}》B3单元格为空，使用sheet名称作为日期")
                            else:
                                # 如果是日期类型，转换为字符串
                                if isinstance(date_value, pd.Timestamp):
                                    date_str = date_value.strftime('%Y-%m-%d')
                                else:
                                    date_str = str(date_value).strip()
                        except:
                            date_str = sheet
                            st.warning(f"⚠️ {file_name} 的 sheet《{sheet}》无法读取B3单元格，使用sheet名称作为日期")
                        
                        # 从第5行开始读取工作内容（索引为4开始，即第5行）
                        # 列为A、B、C、D（索引0、1、2、3）
                        df_work = pd.read_excel(file_path, sheet_name=sheet, header=4, usecols=[0, 1, 2, 3])
                        
                        # 设置列名
                        df_work.columns = ["项目名称", "模块名称", "工作内容", "完成状态"]
                        
                        # 删除所有列都为空的行
                        df_work = df_work.dropna(how='all')
                        
                        # 删除项目名称和模块名称都为空的行
                        df_work = df_work[~(df_work["项目名称"].isna() & df_work["模块名称"].isna())]
                        
                        if df_work.empty:
                            st.warning(f"⚠️ {file_name} 的 sheet《{sheet}》没有有效的工作记录，已跳过。")
                            continue
                        
                        # 添加人员和日期信息
                        df_work["人员"] = person_name
                        df_work["日期"] = date_str
                        
                        file_records.append(df_work)
                        
                        # 调试信息
                        debug_info.append({
                            "文件": file_name,
                            "Sheet": sheet,
                            "人员": person_name,
                            "日期": date_str,
                            "记录数": len(df_work),
                            "状态": "✅ 成功"
                        })
                        
                    except Exception as e:
                        debug_info.append({
                            "文件": file_name,
                            "Sheet": sheet,
                            "人员": "-",
                            "日期": "-",
                            "记录数": 0,
                            "状态": f"❌ 失败: {str(e)}"
                        })
                        st.warning(f"⚠️ 读取 {file_name} 的 sheet《{sheet}》时出错：{e}")
                
                if file_records:
                    return pd.concat(file_records, ignore_index=True)
                    
            except Exception as e:
                st.error(f"❌ 无法读取文件 {file_name}：{e}")
                debug_info.append({
                    "文件": file_name,
                    "Sheet": "-",
                    "人员": "-",
                    "日期": "-",
                    "记录数": 0,
                    "状态": f"❌ 文件错误: {str(e)}"
                })
                
            return pd.DataFrame()

        # === Step 3: 扫描所有 .xlsx 文件 ===
        st.info("📂 正在读取日报文件，请稍候...")
        
        xlsx_files = []
        for root, _, files in os.walk(temp_dir):
            for f in files:
                if f.lower().endswith((".xlsx", ".xls")) and not f.startswith('~$') and not f.startswith('.'):
                    xlsx_files.append(os.path.join(root, f))
        
        if not xlsx_files:
            st.error("❌ 压缩包中没有找到 Excel 文件（.xlsx 或 .xls）")
            st.stop()
        
        st.info(f"找到 {len(xlsx_files)} 个 Excel 文件")
        
        # 读取所有文件
        progress_bar = st.progress(0)
        for idx, fpath in enumerate(xlsx_files):
            df = read_daily_report(fpath)
            if not df.empty:
                all_records.append(df)
            progress_bar.progress((idx + 1) / len(xlsx_files))
        
        # 保存到 session_state
        st.session_state.all_records = all_records
        st.session_state.debug_info = debug_info

        # 显示调试信息
        if debug_info:
            with st.expander("🔍 查看文件读取详情（调试信息）", expanded=True):
                debug_df = pd.DataFrame(debug_info)
                st.dataframe(debug_df, use_container_width=True)

        if not all_records:
            st.error("❌ 未读取到任何有效日报，请检查表格格式。")
            st.error("💡 请确保：")
            st.error("   - B2单元格：人员姓名")
            st.error("   - B3单元格：日期")
            st.error("   - 第5行开始：A列=项目名称、B列=模块名称、C列=工作内容、D列=完成状态")
            st.session_state.processing = False
            st.stop()

        all_data = pd.concat(all_records, ignore_index=True)
        st.success(f"✅ 成功读取 {len(all_data)} 条工作记录，涉及 {all_data['人员'].nunique()} 名人员")
        
        # 保存到 session_state
        st.session_state.all_data = all_data

        # 显示原始数据预览
        with st.expander("📋 查看原始数据", expanded=False):
            st.dataframe(all_data, use_container_width=True)

        # === Step 4: 区分开发 / 测试人员 ===
        tester_names = ["杨妮", "测试"]  # 可在此添加测试人员名单
        all_data["人员类型"] = all_data["人员"].apply(
            lambda x: "测试" if any(t in str(x) for t in tester_names) else "开发"
        )

        # === Step 5: 开发统计 ===
        dev_data = all_data[all_data["人员类型"] == "开发"]
        
        if not dev_data.empty:
            dev_summary = (
                dev_data.groupby(["人员", "模块名称"])
                .size()
                .reset_index(name="维护次数")
            )
            dev_module_count = (
                dev_summary.groupby("人员")["模块名称"].nunique().reset_index(name="模块数量")
            )
            dev_output = pd.merge(dev_module_count, dev_summary, on="人员", how="left")
            
            # 按人员和开发次数排序
            dev_output = dev_output.sort_values(by=["人员", "维护次数"], ascending=[True, False])
        else:
            dev_output = pd.DataFrame(columns=["人员", "模块数量", "模块名称", "维护次数"])
            st.info("ℹ️ 未找到开发人员数据")

        # === Step 6: 测试统计 ===
        test_data = all_data[all_data["人员类型"] == "测试"]
        
        if not test_data.empty:
            test_summary = (
                test_data.groupby(["人员", "模块名称"])
                .size()
                .reset_index(name="测试次数")
            )
            test_module_count = (
                test_summary.groupby("人员")["模块名称"].nunique().reset_index(name="模块数量")
            )
            test_output = pd.merge(test_module_count, test_summary, on="人员", how="left")
            
            # 按人员和测试次数排序
            test_output = test_output.sort_values(by=["人员", "测试次数"], ascending=[True, False])
        else:
            test_output = pd.DataFrame(columns=["人员", "模块数量", "模块名称", "测试次数"])
            st.info("ℹ️ 未找到测试人员数据")
        
        # 保存统计结果到 session_state
        st.session_state.dev_output = dev_output
        st.session_state.test_output = test_output
        st.session_state.results_ready = True
    
    # === 显示结果和下载按钮（独立于处理逻辑） ===
    if hasattr(st.session_state, 'results_ready') and st.session_state.results_ready:
        all_data = st.session_state.all_data
        dev_output = st.session_state.dev_output
        test_output = st.session_state.test_output

        # 显示预览
        st.subheader("📊 统计结果预览")
        
        col1, col2 = st.columns(2)
        with col1:
            st.write(f"**开发人员统计** ({len(dev_output)} 条记录)")
            if not dev_output.empty:
                st.dataframe(dev_output, use_container_width=True)
            else:
                st.info("无数据")
        with col2:
            st.write(f"**测试人员统计** ({len(test_output)} 条记录)")
            if not test_output.empty:
                st.dataframe(test_output, use_container_width=True)
            else:
                st.info("无数据")

        # === Step 7: 输出文件（独立生成，不影响界面） ===
        st.success("🎉 日报处理完成！请下载统计结果👇")
        st.info("💡 提示：可以多次下载")
        
        # 生成下载按钮
        col1, col2, col3 = st.columns([1, 1, 1])
        with col1:
            # 实时生成Excel文件（带格式）
            dev_buffer = BytesIO()
            from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
            
            with pd.ExcelWriter(dev_buffer, engine='openpyxl') as writer:
                dev_output.to_excel(writer, index=False, sheet_name='开发统计')
                workbook = writer.book
                worksheet = writer.sheets['开发统计']
                
                # 设置列宽
                worksheet.column_dimensions['A'].width = 15  # 人员
                worksheet.column_dimensions['B'].width = 12  # 模块数量
                worksheet.column_dimensions['C'].width = 25  # 模块名称
                worksheet.column_dimensions['D'].width = 12  # 维护次数
                
                # 设置表头样式
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=11)
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                for col in range(1, 5):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # 合并单元格并设置样式
                if not dev_output.empty:
                    current_person = None
                    start_row = 2
                    
                    for idx, row in dev_output.iterrows():
                        current_row = idx + 2  # Excel行号（从2开始，1是表头）
                        
                        # 检查是否需要合并上一个人员的单元格
                        if current_person is not None and row['人员'] != current_person:
                            if start_row < current_row:
                                # 合并人员列
                                worksheet.merge_cells(f'A{start_row}:A{current_row - 1}')
                                # 合并模块数量列
                                worksheet.merge_cells(f'B{start_row}:B{current_row - 1}')
                            start_row = current_row
                        
                        current_person = row['人员']
                        
                        # 设置单元格对齐方式
                        for col in range(1, 5):
                            cell = worksheet.cell(row=current_row, column=col)
                            if col in [1, 2, 4]:  # 人员、模块数量、开发次数居中
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            else:  # 模块名称左对齐
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # 处理最后一个人员
                    if start_row < len(dev_output) + 2:
                        worksheet.merge_cells(f'A{start_row}:A{len(dev_output) + 1}')
                        worksheet.merge_cells(f'B{start_row}:B{len(dev_output) + 1}')
                    
                    # 设置合并后单元格的对齐方式
                    for row in range(2, len(dev_output) + 2):
                        worksheet.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                        worksheet.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
            
            dev_buffer.seek(0)
            
            st.download_button(
                label="⬇️ 下载开发人员统计",
                data=dev_buffer.getvalue(),
                file_name="开发人员工作量统计.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_dev"
            )
        
        with col2:
            # 实时生成Excel文件（带格式）
            test_buffer = BytesIO()
            
            with pd.ExcelWriter(test_buffer, engine='openpyxl') as writer:
                test_output.to_excel(writer, index=False, sheet_name='测试统计')
                workbook = writer.book
                worksheet = writer.sheets['测试统计']
                
                # 设置列宽
                worksheet.column_dimensions['A'].width = 15  # 人员
                worksheet.column_dimensions['B'].width = 12  # 模块数量
                worksheet.column_dimensions['C'].width = 25  # 模块名称
                worksheet.column_dimensions['D'].width = 12  # 测试次数
                
                # 设置表头样式
                header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=11)
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                for col in range(1, 5):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # 合并单元格并设置样式
                if not test_output.empty:
                    current_person = None
                    start_row = 2
                    
                    for idx, row in test_output.iterrows():
                        current_row = idx + 2  # Excel行号（从2开始，1是表头）
                        
                        # 检查是否需要合并上一个人员的单元格
                        if current_person is not None and row['人员'] != current_person:
                            if start_row < current_row:
                                # 合并人员列
                                worksheet.merge_cells(f'A{start_row}:A{current_row - 1}')
                                # 合并模块数量列
                                worksheet.merge_cells(f'B{start_row}:B{current_row - 1}')
                            start_row = current_row
                        
                        current_person = row['人员']
                        
                        # 设置单元格对齐方式
                        for col in range(1, 5):
                            cell = worksheet.cell(row=current_row, column=col)
                            if col in [1, 2, 4]:  # 人员、模块数量、测试次数居中
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            else:  # 模块名称左对齐
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # 处理最后一个人员
                    if start_row < len(test_output) + 2:
                        worksheet.merge_cells(f'A{start_row}:A{len(test_output) + 1}')
                        worksheet.merge_cells(f'B{start_row}:B{len(test_output) + 1}')
                    
                    # 设置合并后单元格的对齐方式
                    for row in range(2, len(test_output) + 2):
                        worksheet.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                        worksheet.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
            
            test_buffer.seek(0)
            
            st.download_button(
                label="⬇️ 下载测试人员统计",
                data=test_buffer.getvalue(),
                file_name="测试人员工作量统计.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_test"
            )
        
        with col3:
            # 重新处理按钮
            if st.button("🔄 重新处理", key="reprocess"):
                st.session_state.processing = False
                st.session_state.results_ready = False
                st.rerun()
        
        # 清理临时文件（可选）
        with st.expander("🗑️ 清理临时文件"):
            if st.button("清理临时文件", key="clean_temp"):
                try:
                    if os.path.exists(temp_dir):
                        shutil.rmtree(temp_dir)
                    st.success("✅ 临时文件已清理")
                except Exception as e:
                    st.error(f"清理失败：{e}")
